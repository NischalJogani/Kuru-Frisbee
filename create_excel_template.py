"""
Create a sample Excel template for teams and players upload
Run this to generate a template file you can fill out
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Create Teams sheet
teams_sheet = wb.create_sheet("Teams")
teams_sheet['A1'] = "Team Name"
teams_sheet['A1'].font = Font(bold=True, size=12)
teams_sheet['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
teams_sheet['A1'].font = Font(bold=True, size=12, color="FFFFFF")
teams_sheet['A1'].alignment = Alignment(horizontal="center", vertical="center")

# Add sample team data
sample_teams = ["Team Alpha", "Team Beta", "Team Gamma", "Team Delta", "Team Epsilon", "Team Zeta"]
for idx, team_name in enumerate(sample_teams, start=2):
    teams_sheet[f'A{idx}'] = team_name

teams_sheet.column_dimensions['A'].width = 30

# Create Players sheet
players_sheet = wb.create_sheet("Players")
headers = ["Player Name", "Team", "Jersey Number"]
for col_idx, header in enumerate(headers, start=1):
    cell = players_sheet.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add sample player data (3 players per team)
row = 2
for team in sample_teams:
    for player_num in range(1, 4):
        players_sheet[f'A{row}'] = f"Player {player_num}"
        players_sheet[f'B{row}'] = team
        players_sheet[f'C{row}'] = str(player_num)
        row += 1

# Set column widths
players_sheet.column_dimensions['A'].width = 25
players_sheet.column_dimensions['B'].width = 25
players_sheet.column_dimensions['C'].width = 15

# Save the file
filename = "teams_players_template.xlsx"
wb.save(filename)
print(f"✅ Template created: {filename}")
print("\n📋 Instructions:")
print("1. Open the file in Excel")
print("2. Edit the Teams sheet with your team names")
print("3. Edit the Players sheet with your player data")
print("4. Make sure team names match exactly between sheets")
print("5. Upload the file through the admin panel")
