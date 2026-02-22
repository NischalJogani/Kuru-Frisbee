from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash
from datetime import datetime
from functools import wraps
import json
import os
from werkzeug.utils import secure_filename
from models import db, Admin, Team, Player, Match, Score, TeamSeeding, SpiritScore
import openpyxl

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///frisbee.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
db.init_app(app)

# Create tables and default admin
with app.app_context():
    db.create_all()
    # Create default admin if none exists
    if Admin.query.count() == 0:
        default_admin = Admin(username='admin')
        default_admin.set_password('admin123')
        db.session.add(default_admin)
        db.session.commit()
        print("Default admin created - username: admin, password: admin123")

# Admin authentication decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session:
            flash('Please login to access the admin panel.', 'error')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

# ============ PUBLIC ROUTES ============

@app.route('/leaderboard')
def leaderboard():
    """Public leaderboard with team filtering"""
    teams = Team.query.all()
    selected_team_id = request.args.get('team', type=int)
    
    players = Player.query.all()
    
    # Build leaderboards with rankings
    scoring_data = []
    assist_data = []
    
    for player in players:
        total_points = sum(score.points for score in player.scores if score.action_type == 'score')
        total_assists = player.assists.count()
        
        scoring_data.append({
            'player': player,
            'value': total_points,
            'team': player.team
        })
        
        assist_data.append({
            'player': player,
            'value': total_assists,
            'team': player.team
        })
    
    # Sort all leaderboards
    scoring_data.sort(key=lambda x: -x['value'])
    assist_data.sort(key=lambda x: -x['value'])
    
    # Add global and team ranks
    for idx, entry in enumerate(scoring_data):
        entry['global_rank'] = idx + 1
        entry['team_rank'] = None
    for idx, entry in enumerate(assist_data):
        entry['global_rank'] = idx + 1
        entry['team_rank'] = None
    
    # Calculate team ranks
    for team in teams:
        team_score_players = [e for e in scoring_data if e['team'].id == team.id]
        team_score_players.sort(key=lambda x: -x['value'])
        for idx, entry in enumerate(team_score_players):
            entry['team_rank'] = idx + 1
        
        team_assist_players = [e for e in assist_data if e['team'].id == team.id]
        team_assist_players.sort(key=lambda x: -x['value'])
        for idx, entry in enumerate(team_assist_players):
            entry['team_rank'] = idx + 1
    
    # Filter by selected team
    if selected_team_id:
        scoring_data = [e for e in scoring_data if e['team'].id == selected_team_id]
        assist_data = [e for e in assist_data if e['team'].id == selected_team_id]
    
    return render_template('leaderboard.html',
                         teams=teams,
                         selected_team_id=selected_team_id,
                         scoring_data=scoring_data,
                         assist_data=assist_data)

@app.route('/')
def index():
    """Main public view showing current matches and leaderboard"""
    matches = Match.query.order_by(Match.match_date.desc()).all()
    
    # Get all leaderboards
    players = Player.query.all()
    
    # Scoring leaderboard
    scoring_leaderboard = []
    # Assist leaderboard
    assist_leaderboard = []
    
    for player in players:
        total_points = sum(score.points for score in player.scores if score.action_type == 'score')
        total_assists = player.assists.count()
        
        scoring_leaderboard.append({
            'player': player,
            'total_points': total_points,
            'team': player.team
        })
        
        assist_leaderboard.append({
            'player': player,
            'total_assists': total_assists,
            'team': player.team
        })
    
    scoring_leaderboard.sort(key=lambda x: x['total_points'], reverse=True)
    assist_leaderboard.sort(key=lambda x: x['total_assists'], reverse=True)
    
    return render_template('index.html', 
                         matches=matches, 
                         scoring_leaderboard=scoring_leaderboard[:10],
                         assist_leaderboard=assist_leaderboard[:10])

@app.route('/match/<int:match_id>')
def match_detail(match_id):
    """Live view of a specific match"""
    match = Match.query.get_or_404(match_id)
    scores = Score.query.filter_by(match_id=match_id).order_by(Score.timestamp.desc()).all()
    return render_template('match_detail.html', match=match, scores=scores)

@app.route('/api/match/<int:match_id>/scores')
def get_match_scores(match_id):
    """API endpoint for live score updates"""
    match = Match.query.get_or_404(match_id)
    scores = Score.query.filter_by(match_id=match_id).order_by(Score.timestamp.desc()).all()
    
    # Get current ratios for both teams
    current_ratios = match.get_current_ratio()
    
    return jsonify({
        'team1_score': match.team1_score,
        'team2_score': match.team2_score,
        'status': match.status,
        'current_offense_team_id': match.current_offense_team_id,
        'current_defense_team_id': match.current_defense_team_id,
        'total_points': match.total_points_played,
        'team1_ratio': current_ratios['team1'],
        'team2_ratio': current_ratios['team2'],
        'scores': [{
            'id': score.id,
            'player_name': score.player.name,
            'team_name': score.player.team.name,
            'action_type': score.action_type,
            'points': score.points,
            'timestamp': score.timestamp.strftime('%H:%M:%S'),
            'assist_by': score.assist_by.name if score.assist_by else None
        } for score in scores]
    })

@app.route('/api/match/<int:match_id>/ratio')
def get_match_ratio(match_id):
    """API endpoint for current gender ratio"""
    match = Match.query.get_or_404(match_id)
    return jsonify({
        'ratio': match.get_current_ratio() if match.gender_ratio else None,
        'total_points': match.total_points_played or 0
    })

@app.route('/standings')
def standings():
    """Public standings page with tabs for current, initial, and spirit rankings"""
    teams = Team.query.all()
    completed_matches = Match.query.filter_by(status='completed').all()
    
    # Define stage priority for standings (higher number = higher placement)
    stage_priority = {
        'Finals': 100,           # Winner = 1st, Loser = 2nd
        '3rd Place Game': 50,    # Winner = 3rd, Loser = 4th
        '5th Place Game': 25,    # Winner = 5th, Loser = 6th
        'Cross Pool': 10,        # Regular calculation
        'Pool Stage': 10         # Regular calculation
    }
    
    # Current standings - calculated from match results with stage awareness
    current_standings = []
    
    for team in teams:
        wins = 0
        losses = 0
        point_differential = 0
        final_placement = None
        placement_priority = 0
        
        for match in completed_matches:
            if match.team1_id == team.id:
                score_diff = match.team1_score - match.team2_score
                point_differential += score_diff
                won = match.team1_score > match.team2_score
                
                if won:
                    wins += 1
                else:
                    losses += 1
                
                # Check if this is a placement match
                stage = match.match_stage or 'Pool Stage'
                if stage in ['Finals', '3rd Place Game', '5th Place Game']:
                    if stage == 'Finals':
                        final_placement = 1 if won else 2
                        placement_priority = 200 if won else 199
                    elif stage == '3rd Place Game':
                        final_placement = 3 if won else 4
                        placement_priority = 150 if won else 149
                    elif stage == '5th Place Game':
                        final_placement = 5 if won else 6
                        placement_priority = 100 if won else 99
                        
            elif match.team2_id == team.id:
                score_diff = match.team2_score - match.team1_score
                point_differential += score_diff
                won = match.team2_score > match.team1_score
                
                if won:
                    wins += 1
                else:
                    losses += 1
                
                # Check if this is a placement match
                stage = match.match_stage or 'Pool Stage'
                if stage in ['Finals', '3rd Place Game', '5th Place Game']:
                    if stage == 'Finals':
                        final_placement = 1 if won else 2
                        placement_priority = 200 if won else 199
                    elif stage == '3rd Place Game':
                        final_placement = 3 if won else 4
                        placement_priority = 150 if won else 149
                    elif stage == '5th Place Game':
                        final_placement = 5 if won else 6
                        placement_priority = 100 if won else 99
        
        current_standings.append({
            'team': team,
            'wins': wins,
            'losses': losses,
            'point_diff': point_differential,
            'final_placement': final_placement,
            'placement_priority': placement_priority
        })
    
    # Sort: first by placement priority (if any), then by wins, then by point differential
    current_standings.sort(key=lambda x: (-x['placement_priority'], -x['wins'], -x['point_diff']))
    
    # Initial standings - from seedings
    initial_standings = []
    for team in teams:
        seeding = TeamSeeding.query.filter_by(team_id=team.id).first()
        if seeding:
            initial_standings.append({
                'team': team,
                'seed': seeding.seeding_rank
            })
    initial_standings.sort(key=lambda x: x['seed'])
    
    # Spirit standings - average spirit scores
    spirit_standings = []
    for team in teams:
        spirit_scores_received = SpiritScore.query.filter_by(receiving_team_id=team.id).all()
        
        if spirit_scores_received:
            average_scores = {
                'rules_knowledge': 0,
                'fouls_contact': 0,
                'fair_mindedness': 0,
                'positive_attitude': 0,
                'communication': 0,
                'overall': 0
            }
            
            for score_obj in spirit_scores_received:
                average_scores['rules_knowledge'] += score_obj.rules_knowledge
                average_scores['fouls_contact'] += score_obj.fouls_contact
                average_scores['fair_mindedness'] += score_obj.fair_mindedness
                average_scores['positive_attitude'] += score_obj.positive_attitude
                average_scores['communication'] += score_obj.communication
            
            count = len(spirit_scores_received)
            for key in average_scores:
                if key != 'overall':
                    average_scores[key] = round(average_scores[key] / count, 2)
            
            average_scores['overall'] = round(sum([v for k, v in average_scores.items() if k != 'overall']) / 5, 2)
            
            spirit_standings.append({
                'team': team,
                'match_count': count,
                'scores': average_scores
            })
    
    spirit_standings.sort(key=lambda x: -x['scores']['overall'])
    
    return render_template('standings.html',
                         current_standings=current_standings,
                         initial_standings=initial_standings,
                         spirit_standings=spirit_standings)

@app.route('/spirit-form', methods=['GET', 'POST'])
def spirit_form():
    """Spirit of the Game form page"""
    if request.method == 'POST':
        match_id = request.form.get('match_id')
        giving_team_id = request.form.get('giving_team_id')
        receiving_team_id = request.form.get('receiving_team_id')
        day = request.form.get('day')
        stage = request.form.get('stage')
        
        # Get scores from form
        rules_knowledge = int(request.form.get('rules_knowledge', 3))
        fouls_contact = int(request.form.get('fouls_contact', 3))
        fair_mindedness = int(request.form.get('fair_mindedness', 3))
        positive_attitude = int(request.form.get('positive_attitude', 3))
        communication = int(request.form.get('communication', 3))
        
        mvp_names = request.form.get('mvp_names', '').strip()
        msp_names = request.form.get('msp_names', '').strip()
        feedback = request.form.get('feedback', '').strip()
        
        if match_id and giving_team_id and receiving_team_id:
            spirit_score = SpiritScore(
                match_id=int(match_id),
                giving_team_id=int(giving_team_id),
                receiving_team_id=int(receiving_team_id),
                day=day,
                stage=stage,
                rules_knowledge=rules_knowledge,
                fouls_contact=fouls_contact,
                fair_mindedness=fair_mindedness,
                positive_attitude=positive_attitude,
                communication=communication,
                mvp_names=mvp_names,
                msp_names=msp_names,
                feedback=feedback
            )
            db.session.add(spirit_score)
            db.session.commit()
            flash('Spirit score submitted successfully!', 'success')
            return redirect(url_for('spirit_form'))
    
    teams = Team.query.all()
    completed_matches = Match.query.filter_by(status='completed').all()
    
    return render_template('spirit_form.html', teams=teams, matches=completed_matches)

@app.route('/admin/seeding')
@admin_required
def admin_seeding():
    """Admin page to set tournament seedings"""
    teams = Team.query.all()
    
    # Get existing seedlings
    seedlings = {}
    for team in teams:
        seeding = TeamSeeding.query.filter_by(team_id=team.id).first()
        seedlings[team.id] = seeding.seeding_rank if seeding else None
    
    return render_template('admin/seeding.html', teams=teams, seedlings=seedlings)

@app.route('/admin/seeding/update', methods=['POST'])
@admin_required
def update_seeding():
    """Update team seedings"""
    teams_data = request.get_json()
    
    for team_id, seed_rank in teams_data.items():
        team_id = int(team_id)
        seeding = TeamSeeding.query.filter_by(team_id=team_id).first()
        
        if seeding and seed_rank:
            seeding.seeding_rank = int(seed_rank)
        elif seed_rank and not seeding:
            new_seeding = TeamSeeding(team_id=team_id, seeding_rank=int(seed_rank))
            db.session.add(new_seeding)
        elif seeding and not seed_rank:
            db.session.delete(seeding)
    
    db.session.commit()
    return jsonify({'status': 'success'})

@app.route('/api/standings/spirit')
def get_spirit_standings_api():
    """API endpoint for spirit standings (for updates)"""
    teams = Team.query.all()
    spirit_standings = []
    
    for team in teams:
        spirit_scores_received = SpiritScore.query.filter_by(receiving_team_id=team.id).all()
        
        if spirit_scores_received:
            average_scores = {
                'rules_knowledge': 0,
                'fouls_contact': 0,
                'fair_mindedness': 0,
                'positive_attitude': 0,
                'communication': 0
            }
            
            for score_obj in spirit_scores_received:
                average_scores['rules_knowledge'] += score_obj.rules_knowledge
                average_scores['fouls_contact'] += score_obj.fouls_contact
                average_scores['fair_mindedness'] += score_obj.fair_mindedness
                average_scores['positive_attitude'] += score_obj.positive_attitude
                average_scores['communication'] += score_obj.communication
            
            count = len(spirit_scores_received)
            for key in average_scores:
                average_scores[key] = round(average_scores[key] / count, 2)
            
            overall = round(sum(average_scores.values()) / 5, 2)
            
            spirit_standings.append({
                'team_id': team.id,
                'team_name': team.name,
                'match_count': count,
                'overall': overall,
                'scores': average_scores
            })
    
    spirit_standings.sort(key=lambda x: -x['overall'])
    return jsonify(spirit_standings)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if 'admin_id' in session:
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        admin = Admin.query.filter_by(username=username).first()
        if admin and admin.check_password(password):
            session['admin_id'] = admin.id
            session['admin_username'] = admin.username
            flash('Login successful!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
def admin_logout():
    """Admin logout"""
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    flash('Logged out successfully', 'success')
    return redirect(url_for('index'))

# ============ ADMIN ROUTES ============

@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard home"""
    teams_count = Team.query.count()
    players_count = Player.query.count()
    matches_count = Match.query.count()
    live_matches = Match.query.filter_by(status='live').all()
    
    return render_template('admin/dashboard.html', 
                         teams_count=teams_count,
                         players_count=players_count,
                         matches_count=matches_count,
                         live_matches=live_matches)

# --- TEAM MANAGEMENT ---

@app.route('/admin/teams')
@admin_required
def admin_teams():
    """View all teams"""
    teams = Team.query.all()
    return render_template('admin/teams.html', teams=teams)

@app.route('/admin/teams/add', methods=['POST'])
@admin_required
def add_team():
    """Add a new team"""
    name = request.form.get('name')
    if name:
        team = Team(name=name)
        db.session.add(team)
        db.session.commit()
    return redirect(url_for('admin_teams'))

@app.route('/admin/teams/delete/<int:team_id>', methods=['POST'])
@admin_required
def delete_team(team_id):
    """Delete a team"""
    team = Team.query.get_or_404(team_id)
    
    # Delete associated seeding record if it exists
    seeding = TeamSeeding.query.filter_by(team_id=team_id).first()
    if seeding:
        db.session.delete(seeding)
    
    # Delete the team (will cascade to players due to model relationship)
    db.session.delete(team)
    db.session.commit()
    return redirect(url_for('admin_teams'))

# --- PLAYER MANAGEMENT ---

@app.route('/admin/players')
@admin_required
def admin_players():
    """View all players"""
    players = Player.query.all()
    teams = Team.query.all()
    return render_template('admin/players.html', players=players, teams=teams)

@app.route('/admin/players/add', methods=['POST'])
@admin_required
def add_player():
    """Add a new player"""
    name = request.form.get('name')
    team_id = request.form.get('team_id')
    jersey_number = request.form.get('jersey_number')
    
    if name and team_id:
        player = Player(name=name, team_id=team_id, jersey_number=jersey_number)
        db.session.add(player)
        db.session.commit()
    return redirect(url_for('admin_players'))

@app.route('/admin/players/delete/<int:player_id>', methods=['POST'])
@admin_required
def delete_player(player_id):
    """Delete a player"""
    player = Player.query.get_or_404(player_id)
    db.session.delete(player)
    db.session.commit()
    return redirect(url_for('admin_players'))

@app.route('/admin/upload-excel', methods=['GET', 'POST'])
@admin_required
def upload_excel():
    """Upload Excel file to bulk import teams and players"""
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file uploaded', 'error')
            return redirect(request.url)
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'error')
            return redirect(request.url)
        
        try:
            # Load the Excel file
            wb = openpyxl.load_workbook(file)
            
            teams_added = 0
            players_added = 0
            errors = []
            
            # Process Teams sheet
            if 'Teams' in wb.sheetnames or 'teams' in wb.sheetnames:
                teams_sheet = wb['Teams'] if 'Teams' in wb.sheetnames else wb['teams']
                
                # Get existing team names to avoid duplicates
                existing_teams = {team.name.lower(): team for team in Team.query.all()}
                team_name_to_id = {}
                
                # Start from row 2 to skip header
                for row in teams_sheet.iter_rows(min_row=2, values_only=True):
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    team_name = str(row[0]).strip()
                    
                    if not team_name:
                        continue
                    
                    # Check if team already exists
                    if team_name.lower() in existing_teams:
                        team = existing_teams[team_name.lower()]
                        team_name_to_id[team_name] = team.id
                        errors.append(f"Team '{team_name}' already exists - skipped")
                    else:
                        # Create new team
                        team = Team(name=team_name)
                        db.session.add(team)
                        db.session.flush()  # Get the ID
                        team_name_to_id[team_name] = team.id
                        existing_teams[team_name.lower()] = team
                        teams_added += 1
                
                db.session.commit()
            else:
                errors.append("No 'Teams' sheet found in Excel file")
            
            # Process Players sheet
            if 'Players' in wb.sheetnames or 'players' in wb.sheetnames:
                players_sheet = wb['Players'] if 'Players' in wb.sheetnames else wb['players']
                
                # Reload teams after commit
                existing_teams = {team.name.lower(): team for team in Team.query.all()}
                team_name_to_id = {team.name: team.id for team in Team.query.all()}
                
                # Start from row 2 to skip header
                for row_num, row in enumerate(players_sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    player_name = str(row[0]).strip() if row[0] else None
                    team_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                    jersey_number = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    
                    if not player_name:
                        errors.append(f"Row {row_num}: Missing player name - skipped")
                        continue
                    
                    if not team_name:
                        errors.append(f"Row {row_num}: Missing team for player '{player_name}' - skipped")
                        continue
                    
                    # Find team
                    if team_name.lower() not in existing_teams:
                        errors.append(f"Row {row_num}: Team '{team_name}' not found for player '{player_name}' - skipped")
                        continue
                    
                    team = existing_teams[team_name.lower()]
                    
                    # Create player
                    player = Player(
                        name=player_name,
                        team_id=team.id,
                        jersey_number=jersey_number
                    )
                    db.session.add(player)
                    players_added += 1
                
                db.session.commit()
            else:
                errors.append("No 'Players' sheet found in Excel file")
            
            # Success message
            success_msg = f"Successfully added {teams_added} teams and {players_added} players!"
            flash(success_msg, 'success')
            
            if errors:
                flash(f"Warnings: {len(errors)} issues encountered", 'warning')
                for error in errors[:10]:  # Show first 10 errors
                    flash(error, 'warning')
            
            return redirect(url_for('admin_players'))
            
        except Exception as e:
            flash(f'Error processing Excel file: {str(e)}', 'error')
            return redirect(request.url)
    
    # GET request - show upload form
    teams_count = Team.query.count()
    players_count = Player.query.count()
    return render_template('admin/upload_excel.html', 
                         teams_count=teams_count,
                         players_count=players_count)


# --- MATCH MANAGEMENT ---

@app.route('/admin/matches')
@admin_required
def admin_matches():
    """View all matches"""
    matches = Match.query.order_by(Match.match_date.desc()).all()
    teams = Team.query.all()
    return render_template('admin/matches.html', matches=matches, teams=teams)

@app.route('/admin/matches/add', methods=['POST'])
@admin_required
def add_match():
    """Add a new match"""
    team1_id = request.form.get('team1_id')
    team2_id = request.form.get('team2_id')
    match_date = request.form.get('match_date')
    location = request.form.get('location')
    duration_minutes = request.form.get('duration_minutes', '60')
    max_score = request.form.get('max_score', '15')
    match_stage = request.form.get('match_stage', 'Pool Stage')
    
    if team1_id and team2_id and match_date:
        match = Match(
            team1_id=team1_id,
            team2_id=team2_id,
            match_date=datetime.fromisoformat(match_date),
            location=location,
            duration_minutes=int(duration_minutes),
            max_score=int(max_score),
            match_stage=match_stage,
            status='scheduled'
        )
        db.session.add(match)
        db.session.commit()
    return redirect(url_for('admin_matches'))

@app.route('/admin/matches/update_status/<int:match_id>', methods=['POST'])
@admin_required
def update_match_status(match_id):
    """Update match status"""
    match = Match.query.get_or_404(match_id)
    status = request.form.get('status')
    if status:
        match.status = status
        db.session.commit()
    return redirect(url_for('admin_matches'))

@app.route('/admin/matches/delete/<int:match_id>', methods=['POST'])
@admin_required
def delete_match(match_id):
    """Delete a match"""
    match = Match.query.get_or_404(match_id)
    db.session.delete(match)
    db.session.commit()
    return redirect(url_for('admin_matches'))

# --- LIVE SCORING ---

@app.route('/admin/scoring/<int:match_id>')
@admin_required
def admin_scoring(match_id):
    """Live scoring interface for a match"""
    match = Match.query.get_or_404(match_id)
    team1_players = Player.query.filter_by(team_id=match.team1_id).all()
    team2_players = Player.query.filter_by(team_id=match.team2_id).all()
    scores = Score.query.filter_by(match_id=match_id).order_by(Score.timestamp.desc()).all()
    
    # Convert players to dictionaries for JSON serialization
    team1_players_dict = [p.to_dict() for p in team1_players]
    team2_players_dict = [p.to_dict() for p in team2_players]
    
    return render_template('admin/live_scoring.html', 
                         match=match,
                         team1_players=team1_players,
                         team2_players=team2_players,
                         team1_players_dict=team1_players_dict,
                         team2_players_dict=team2_players_dict,
                         scores=scores)

@app.route('/admin/scoring/<int:match_id>/add', methods=['POST'])
@admin_required
def add_action(match_id):
    """Add a score or defense action to a match"""
    match = Match.query.get_or_404(match_id)
    player_id = request.form.get('player_id')
    action_type = request.form.get('action_type')  # 'score' or 'defense'
    
    if player_id and action_type:
        if action_type == 'score':
            points = int(request.form.get('points', 1))
            assist_player_id = request.form.get('assist_player_id') or None
            
            # Create score record
            score = Score(
                match_id=match_id,
                player_id=player_id,
                action_type='score',
                points=points,
                assist_player_id=assist_player_id,
                timestamp=datetime.now()
            )
            db.session.add(score)
            
            # Update match score
            player = Player.query.get(player_id)
            if player.team_id == match.team1_id:
                match.team1_score += points
            else:
                match.team2_score += points
            
            # Switch possession - scoring team now on defense
            if player.team_id == match.current_offense_team_id:
                # Offense team just scored, so they become defense
                match.current_offense_team_id = match.current_defense_team_id
                match.current_defense_team_id = player.team_id
            
            # Increment total points played for ratio tracking
            match.total_points_played += 1
            
            # Check if match has reached max score
            if match.team1_score >= match.max_score or match.team2_score >= match.max_score:
                match.status = 'completed'
                flash('Match completed!', 'success')
        
        db.session.commit()
    
    return redirect(url_for('admin_scoring', match_id=match_id))

@app.route('/admin/scoring/<int:match_id>/undo/<int:score_id>', methods=['POST'])
@admin_required
def undo_action(match_id, score_id):
    """Undo a score or defense action"""
    match = Match.query.get_or_404(match_id)
    score = Score.query.get_or_404(score_id)
    
    # Update match score only for actual scoring actions (not defense)
    if score.action_type == 'score':
        player = Player.query.get(score.player_id)
        if player.team_id == match.team1_id:
            match.team1_score -= score.points
        else:
            match.team2_score -= score.points
    
    db.session.delete(score)
    db.session.commit()
    
    return redirect(url_for('admin_scoring', match_id=match_id))
    
    db.session.delete(score)
    db.session.commit()
    
    return redirect(url_for('admin_scoring', match_id=match_id))

@app.route('/admin/scoring/<int:match_id>/start', methods=['POST'])
@admin_required
def start_match(match_id):
    """Start a match and set initial offense/defense and gender ratio"""
    match = Match.query.get_or_404(match_id)
    offense_team_id = request.form.get('offense_team_id')
    gender_ratio = request.form.get('gender_ratio')
    
    if offense_team_id:
        match.start_time = datetime.now()
        match.status = 'live'
        match.current_offense_team_id = int(offense_team_id)
        match.gender_ratio = gender_ratio
        match.total_points_played = 0
        # Defense is the other team
        if int(offense_team_id) == match.team1_id:
            match.current_defense_team_id = match.team2_id
        else:
            match.current_defense_team_id = match.team1_id
        db.session.commit()
        flash(f'Match started!', 'success')
    
    return redirect(url_for('admin_scoring', match_id=match_id))

@app.route('/admin/scoring/<int:match_id>/set-possession', methods=['POST'])
@admin_required
def set_possession(match_id):
    """Set or switch offense/defense during a match"""
    match = Match.query.get_or_404(match_id)
    offense_team_id = request.form.get('offense_team_id')
    
    if offense_team_id:
        match.current_offense_team_id = int(offense_team_id)
        if int(offense_team_id) == match.team1_id:
            match.current_defense_team_id = match.team2_id
        else:
            match.current_defense_team_id = match.team1_id
        db.session.commit()
        flash('Possession updated!', 'success')
    
    return redirect(url_for('admin_scoring', match_id=match_id))

@app.route('/admin/scoring/<int:match_id>/set-ratio', methods=['POST'])
@admin_required
def set_ratio(match_id):
    """Update gender ratio during a match"""
    match = Match.query.get_or_404(match_id)
    gender_ratio = request.form.get('gender_ratio')
    
    if gender_ratio:
        match.gender_ratio = gender_ratio
        db.session.commit()
        flash('Ratio updated!', 'success')
    
    return redirect(url_for('admin_scoring', match_id=match_id))

@app.route('/admin/scoring/<int:match_id>/end', methods=['POST'])
@admin_required
def end_match(match_id):
    """End a match manually"""
    match = Match.query.get_or_404(match_id)
    match.status = 'completed'
    db.session.commit()
    flash('Match ended!', 'success')
    return redirect(url_for('admin_scoring', match_id=match_id))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5011)
